"""Export the Joe Knows Ball 2026 fantasy rankings workbook into src/data/fantasyRankings2026.ts.

Source of truth: Fantasy_Football_Rankings_V2_Complete (1).xlsx
  - "Main Rankings" tab rows 5-254 are the 250 ranked players (columns A-R).
  - "Team Context" tab rows 2-251 provide canonical team identity, aligned 1:1
    with the Main Rankings rows (verified by this script before emitting).
  - Rounds/picks come from columns A (round, merged cell, forward-filled) and B
    (pick within round). Rounds 1-20 have 12 picks; round 21 has 10.

The emitted rows are a faithful transcription: overallRank is the workbook row
order (1-250), every rank value is parsed from the workbook verbatim, and blank
cells stay undefined in the output (never coerced to zero). Team codes are
normalized to the site's lowercase canonical codes via Team Context; free agents
("FA") get no team field.

Run from the repo root:
  python scripts/export-fantasy-rankings-2026.py [workbook path]
"""

import json
import re
import sys

WORKBOOK_DEFAULT = r"C:\Users\jbloo\Downloads\Fantasy_Football_Rankings_V2_Complete (1).xlsx"
OUTPUT = "src/data/fantasyRankings2026.ts"

# Workbook Team Context code -> site canonical lowercase code.
TEAM_MAP = {
    "ARI": "ari", "ATL": "atl", "BAL": "bal", "BUF": "buf", "CAR": "car",
    "CHI": "chi", "CIN": "cin", "CLE": "cle", "DAL": "dal", "DEN": "den",
    "DET": "det", "GB": "gb", "HOU": "hou", "IND": "ind", "JAX": "jax",
    "KC": "kc", "LAC": "lac", "LAR": "lar", "LV": "lv", "MIA": "mia",
    "MIN": "min", "NE": "ne", "NO": "no", "NYG": "nyg", "NYJ": "nyj",
    "PHI": "phi", "PIT": "pit", "SEA": "sea", "SF": "sf", "TB": "tb",
    "TEN": "ten", "WAS": "wsh",
}

# Position-specific metric label -> row field name, in workbook column order.
METRIC_FIELDS = {
    "QB": [
        ("passerRatingRank", "Passer Rating"),
        ("rushingYardsPerGameRank", "Rush Yds/Game"),
        ("passTdPerAttemptRank", "Pass TD/Attempt"),
    ],
    "RB": [
        ("touchesRank", "Touches"),
        ("redZoneTouchesRank", "Red Zone Touches"),
        ("ypcRank", "YPC"),
    ],
    "WR": [
        ("targetPercentRank", "Target Percent"),
        ("airYardsPerGameRank", "Air Yards/Game"),
        ("targetsPerGameRank", "Targets/Game"),
    ],
    "TE": [
        ("targetShareRank", "Target Share"),
        ("targetsPerRouteRunRank", "Targets/Route Run"),
        ("yprrRank", "YPRR"),
    ],
}

RANK_RE = re.compile(r"^(QB|RB|WR|TE)(\d+)$")


def parse_rank(value, position):
    """Parse a 'RB7'-style rank cell into an int, or None if blank."""
    if value is None or value == "":
        return None
    match = RANK_RE.match(str(value).strip())
    if not match:
        raise ValueError(f"unexpected rank cell {value!r} (position {position})")
    if match.group(1) != position:
        raise ValueError(
            f"rank cell {value!r} prefix {match.group(1)!r} != row position {position!r}"
        )
    return int(match.group(2))


def load_workbook(path):
    import openpyxl

    return openpyxl.load_workbook(path, data_only=True)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else WORKBOOK_DEFAULT
    wb = load_workbook(path)
    main = wb["Main Rankings"]
    teams = wb["Team Context"]

    # --- Read Team Context (team identity + free-agent status) ---
    team_rows = []
    for i in range(250):
        row = 2 + i
        name = teams.cell(row=row, column=1).value
        code = teams.cell(row=row, column=2).value
        pos = teams.cell(row=row, column=3).value
        team_rows.append({"name": name, "code": code, "pos": pos})

    # --- Read Main Rankings rows 5-254 ---
    rows = []
    current_round = None
    for i in range(250):
        r = 5 + i
        cell_round = main.cell(row=r, column=1).value
        if cell_round is not None:
            current_round = int(cell_round)
        pick = main.cell(row=r, column=2).value
        player = main.cell(row=r, column=3).value
        pos_rank = main.cell(row=r, column=4).value
        position = main.cell(row=r, column=5).value

        if position not in METRIC_FIELDS:
            raise ValueError(f"row {r}: unexpected position {position!r}")
        if current_round is None or pick is None or player is None:
            raise ValueError(f"row {r}: missing round/pick/player")

        tc = team_rows[i]
        if tc["name"] != player:
            raise ValueError(
                f"row {r}: Team Context player {tc['name']!r} != Main Rankings {player!r}"
            )
        if tc["pos"] != position:
            raise ValueError(
                f"row {r}: Team Context pos {tc['pos']!r} != Main Rankings {position!r}"
            )

        team = TEAM_MAP.get(tc["code"])
        if team is None and tc["code"] != "FA":
            raise ValueError(f"row {r}: unknown team code {tc['code']!r}")
        # Team Context must stay blank-only for free agents.
        if tc["code"] == "FA":
            for col in (4, 5, 6, 7, 8, 9):
                if teams.cell(row=2 + i, column=col).value not in (None, ""):
                    raise ValueError(f"row {r}: FA player has Team Context data")

        pos_rank_n = parse_rank(pos_rank, position)

        war = parse_rank(main.cell(row=r, column=9).value, position)
        last8 = parse_rank(main.cell(row=r, column=10).value, position)
        proj = parse_rank(main.cell(row=r, column=11).value, position)
        vegas = parse_rank(main.cell(row=r, column=12).value, position)
        avg = parse_rank(main.cell(row=r, column=13).value, position)

        sos = main.cell(row=r, column=14).value
        oline = main.cell(row=r, column=15).value
        w15 = main.cell(row=r, column=16).value
        w16 = main.cell(row=r, column=17).value
        w17 = main.cell(row=r, column=18).value

        metrics = {}
        for idx, (field, _label) in enumerate(METRIC_FIELDS[position]):
            val = parse_rank(main.cell(row=r, column=6 + idx).value, position)
            if val is not None:
                metrics[field] = val

        row_obj = {
            "overallRank": i + 1,
            "player": player,
            "position": position,
            "positionRank": pos_rank_n,
            "draftRound": current_round,
            "roundPick": int(pick),
            "averageRank": avg,
        }
        if team is not None:
            row_obj["team"] = team
        if war is not None:
            row_obj["warRank"] = war
        if last8 is not None:
            row_obj["lateSeasonRank"] = last8
        if proj is not None:
            row_obj["projectionRank"] = proj
        if vegas is not None:
            row_obj["vegasRank"] = vegas
        if sos is not None:
            row_obj["strengthOfSchedule"] = float(sos)
        if oline is not None:
            row_obj["offensiveLineRank"] = float(oline)
        if w15 is not None:
            row_obj["playoffWeek15Opponent"] = str(w15)
        if w16 is not None:
            row_obj["playoffWeek16Opponent"] = str(w16)
        if w17 is not None:
            row_obj["playoffWeek17Opponent"] = str(w17)
        if metrics:
            row_obj["metrics"] = metrics

        rows.append(row_obj)

    # --- Verification before emitting ---
    assert len(rows) == 250, len(rows)
    ranks = [row["overallRank"] for row in rows]
    assert ranks == list(range(1, 251)), "overallRank must be sequential 1-250"
    round21 = [row for row in rows if row["draftRound"] == 21]
    assert len(round21) == 10, len(round21)
    for row in round21:
        assert 1 <= row["roundPick"] <= 10, row
    for row in rows:
        assert 1 <= row["draftRound"] <= 21, row
        assert 1 <= row["roundPick"] <= 12, row

    # Emit TS.
    body = "\n".join("  " + json.dumps(row, ensure_ascii=False) + "," for row in rows)
    ts = f"""// GENERATED — do not hand-edit.
// Source: Fantasy_Football_Rankings_V2_Complete (1).xlsx (Main Rankings rows 5-254,
// Team Context rows 2-251). Extracted verbatim by scripts/export-fantasy-rankings-2026.py.
// - overallRank is workbook order (1-250). Nothing is reordered or recalculated.
// - Blank workbook cells stay undefined; never coerced to zero.
// - team uses the site's lowercase canonical codes; free agents carry no team.
// - Per-position metrics (metrics) are the workbook's own rank values.
import type {{ FantasyRankingRow }} from "@/lib/fantasy/rankings";

export const FANTASY_RANKING_ROWS_2026: readonly FantasyRankingRow[] = [
{body}
];
"""
    with open(OUTPUT, "w", encoding="utf-8", newline="\n") as fh:
        fh.write(ts)
    print(f"Wrote {OUTPUT}: {len(rows)} rows")
    print("First:", rows[0]["player"], rows[0]["team"], rows[0]["positionRank"])
    print("Last:", rows[-1]["player"], rows[-1]["team"], rows[-1]["positionRank"])


if __name__ == "__main__":
    main()
