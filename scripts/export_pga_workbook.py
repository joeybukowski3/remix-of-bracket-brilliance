from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


STAT_FIELDS = [
    "SG: Approach the Green",
    "SG: Around the Green",
    "SG: Putting",
    "Par 4 Scoring Average",
    "Driving Accuracy %",
    "Bogey Avoidance",
    "Birdie or Better 125-150 yds",
    "Birdie or Better <125 yds",
]

STAT_DIRECTIONS = {
    "SG: Approach the Green": "desc",
    "SG: Around the Green": "desc",
    "SG: Putting": "desc",
    "Par 4 Scoring Average": "asc",
    "Driving Accuracy %": "desc",
    "Bogey Avoidance": "desc",
    "Birdie or Better 125-150 yds": "desc",
    "Birdie or Better <125 yds": "desc",
}

NAME_ALIASES = {
    "matthew mccarty": ["matt mccarty"],
    "jordan l smith": ["jordan smith"],
    "nicolas echavarria": ["nico echavarria"],
    "john keefer": ["johnny keefer"],
}


@dataclass
class JoinResult:
    values: dict[str, Any] | None
    candidate_keys: list[str]


def normalize_name(name: str | None) -> str:
    if not name:
        return ""

    text = str(name)
    for original, replacement in {
        "ø": "o",
        "Ø": "O",
        "å": "a",
        "Å": "A",
        "æ": "ae",
        "Æ": "AE",
        "ö": "o",
        "Ö": "O",
        "ä": "a",
        "Ä": "A",
    }.items():
        text = text.replace(original, replacement)

    ascii_name = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    ascii_name = ascii_name.lower().replace("&", " and ")
    ascii_name = re.sub(r"[^\w\s]", " ", ascii_name)
    tokens = [token for token in ascii_name.split() if token]
    tokens = [token for token in tokens if not (len(token) == 1 and token.isalpha())]
    return " ".join(tokens)


def build_candidate_keys(name: str) -> list[str]:
    normalized = normalize_name(name)
    candidates = [normalized]
    for alias in NAME_ALIASES.get(normalized, []):
        alias_key = normalize_name(alias)
        if alias_key and alias_key not in candidates:
            candidates.append(alias_key)
    return candidates


def parse_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, (int, float)):
        if isinstance(value, float) and not math.isfinite(value):
            return None
        return value

    text = str(value).strip()
    if not text or text.lower() in {"null", "nan", "none"}:
        return None

    text = text.replace(",", "")
    if text.endswith("%"):
        text = text[:-1]

    try:
        return float(text)
    except ValueError:
        return None


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"null", "nan", "none"}:
        return None
    return text


def convert_last_first_name(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if "," not in text:
        return normalize_text(text)
    last, first = [part.strip() for part in text.split(",", 1)]
    return normalize_text(f"{first} {last}".strip())


def load_workbook(path: Path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def iter_sheet_rows(workbook, sheet_name: str, min_row: int):
    sheet = workbook[sheet_name]
    yield from sheet.iter_rows(min_row=min_row, values_only=True)


def build_index(rows: list[dict[str, Any]], name_key: str = "Player Name") -> tuple[dict[str, dict[str, Any]], dict[str, list[str]]]:
    index: dict[str, dict[str, Any]] = {}
    duplicates: dict[str, list[str]] = {}
    for row in rows:
        key = normalize_name(str(row[name_key]))
        if not key:
            continue
        if key in index:
            duplicates.setdefault(key, [str(index[key][name_key])]).append(str(row[name_key]))
            continue
        index[key] = row
    return index, duplicates


def resolve_join(name: str, index: dict[str, dict[str, Any]]) -> JoinResult:
    candidates = build_candidate_keys(name)
    for candidate in candidates:
        if candidate in index:
            return JoinResult(index[candidate], candidates)
    return JoinResult(None, candidates)


def load_base_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in iter_sheet_rows(workbook, sheet_name, 4):
        if not row or row[4] is None or row[5] is None:
            continue
        rows.append(
            {
                "Player Name": str(row[5]).strip(),
                "Salary": parse_number(row[6]),
                "DK Avg PPG": parse_number(row[7]),
                "Adj Proj Score": parse_number(row[9]),
                "Adj Value": parse_number(row[10]),
            }
        )
    return rows


def build_stats_only_base_rows(stat_rows: list[dict[str, Any]], trend_index: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for stat_row in stat_rows:
        player_name = str(stat_row["Player Name"]).strip()
        trend = resolve_join(player_name, trend_index)
        rows.append(
            {
                "Player Name": player_name,
                "Salary": None,
                "DK Avg PPG": None,
                "Adj Proj Score": trend.values.get("TrendRank") if trend.values else None,
                "Adj Value": None,
            }
        )
    return rows


def load_trend_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in iter_sheet_rows(workbook, sheet_name, 2):
        if not row or row[0] is None:
            continue
        display_name = row[10] if len(row) > 10 and row[10] else convert_last_first_name(row[0])
        if not display_name:
            continue
        rows.append({"Player Name": display_name, "TrendRank": parse_number(row[2])})
    return rows


def load_history_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in iter_sheet_rows(workbook, sheet_name, 2):
        if not row or row[0] is None:
            continue
        display_name = row[12] if len(row) > 12 and row[12] else convert_last_first_name(row[0])
        if not display_name:
            continue
        rows.append(
            {
                "Player Name": str(display_name).strip(),
                "2021": normalize_text(row[1]),
                "2022": normalize_text(row[2]),
                "2023": normalize_text(row[3]),
                "2024": normalize_text(row[4]),
                "2025": normalize_text(row[5]),
                "HT # Rounds": parse_number(row[6]),
                "Course True SG": parse_number(row[7]),
            }
        )
    return rows


def load_stat_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in iter_sheet_rows(workbook, sheet_name, 3):
        if not row or row[0] is None:
            continue
        rows.append(
            {
                "Player Name": str(row[0]).strip(),
                "SG: Approach the Green": parse_number(row[4]),
                "SG: Around the Green": parse_number(row[5]),
                "SG: Putting": parse_number(row[6]),
                "Par 4 Scoring Average": parse_number(row[8]),
                "Driving Accuracy %": parse_number(row[20]),
                "Bogey Avoidance": parse_number(row[18]),
                "Birdie or Better 125-150 yds": parse_number(row[13]),
                "Birdie or Better <125 yds": parse_number(row[14]),
            }
        )
    return rows


def compute_stat_ranks(players: list[dict[str, Any]]):
    for stat in STAT_FIELDS:
        populated = [player for player in players if player.get(stat) is not None]
        if STAT_DIRECTIONS[stat] == "asc":
            populated.sort(key=lambda player: (float(player[stat]), player["Player Name"]))
        else:
            populated.sort(key=lambda player: (-float(player[stat]), player["Player Name"]))

        previous_value: float | None = None
        previous_rank = 0
        for position, player in enumerate(populated, start=1):
            numeric_value = float(player[stat])
            rank = previous_rank if previous_value is not None and numeric_value == previous_value else position
            player[f"{stat}_rank"] = rank
            previous_rank = rank
            previous_value = numeric_value

        for player in players:
            if player.get(stat) is None:
                player[f"{stat}_rank"] = None


def build_player_export(
    base_row: dict[str, Any],
    trend_index: dict[str, dict[str, Any]],
    history_index: dict[str, dict[str, Any]],
    stat_index: dict[str, dict[str, Any]],
    warnings: list[str],
) -> dict[str, Any]:
    player_name = str(base_row["Player Name"])
    trend = resolve_join(player_name, trend_index)
    history = resolve_join(player_name, history_index)
    stats = resolve_join(player_name, stat_index)

    if trend.values is None:
        warnings.append(f"Trend join missing for {player_name} (candidates: {', '.join(trend.candidate_keys)})")
    if history.values is None:
        warnings.append(f"History join missing for {player_name} (candidates: {', '.join(history.candidate_keys)})")
    if stats.values is None:
        warnings.append(f"Stat join missing for {player_name} (candidates: {', '.join(stats.candidate_keys)})")

    player = {
        "Player Name": player_name,
        "Salary": base_row.get("Salary"),
        "DK Avg PPG": base_row.get("DK Avg PPG"),
        "HT # Rounds": history.values.get("HT # Rounds") if history.values else None,
        "Course True SG": history.values.get("Course True SG") if history.values else None,
        "2021": history.values.get("2021") if history.values else None,
        "2022": history.values.get("2022") if history.values else None,
        "2023": history.values.get("2023") if history.values else None,
        "2024": history.values.get("2024") if history.values else None,
        "2025": history.values.get("2025") if history.values else None,
        "TrendRank": trend.values.get("TrendRank") if trend.values else None,
        "Adj Proj Score": base_row.get("Adj Proj Score"),
        "Adj Value": base_row.get("Adj Value"),
        "Model Rank": None,
    }

    for stat in STAT_FIELDS:
        player[stat] = stats.values.get(stat) if stats.values else None

    available_stat_count = sum(1 for stat in STAT_FIELDS if player.get(stat) is not None)
    player["hasStatProfile"] = available_stat_count > 0
    player["missingStatFields"] = [stat for stat in STAT_FIELDS if player.get(stat) is None]
    player["dataCompletenessScore"] = round(available_stat_count / len(STAT_FIELDS), 3)
    return player


def write_json(path: Path, payload: list[dict[str, Any]]):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export a PGA tournament JSON file from a workbook.")
    parser.add_argument("--workbook", required=True, type=Path, help="Workbook containing DK/base, trend, history, and stat sheets.")
    parser.add_argument("--output", required=True, type=Path, action="append", help="Output JSON path. Pass multiple times to mirror outputs.")
    parser.add_argument("--base-sheet", default="DK_RBC")
    parser.add_argument("--trend-sheet", default="TrendTable")
    parser.add_argument("--history-sheet", default="HarbourTownHistory")
    parser.add_argument("--stats-sheet", default="PGA_Stats_Master")
    parser.add_argument("--base-mode", choices=("sheet", "stats"), default="sheet")
    parser.add_argument("--strict-missing-stats", action="store_true")
    args = parser.parse_args()

    workbook = load_workbook(args.workbook)
    trend_rows = load_trend_rows(workbook, args.trend_sheet)
    stat_rows = load_stat_rows(workbook, args.stats_sheet)

    trend_index, trend_duplicates = build_index(trend_rows)
    stat_index, stat_duplicates = build_index(stat_rows)

    if args.base_mode == "stats":
        base_rows = build_stats_only_base_rows(stat_rows, trend_index)
        history_rows: list[dict[str, Any]] = []
    else:
        base_rows = load_base_rows(workbook, args.base_sheet)
        history_rows = load_history_rows(workbook, args.history_sheet)

    history_index, history_duplicates = build_index(history_rows)

    warnings: list[str] = []
    for label, duplicates in [("trend", trend_duplicates), ("history", history_duplicates), ("stats", stat_duplicates)]:
        for key, names in duplicates.items():
            warnings.append(f"Duplicate {label} key '{key}' for {', '.join(names)}")

    players = [
        build_player_export(base_row, trend_index, history_index, stat_index, warnings)
        for base_row in base_rows
    ]

    compute_stat_ranks(players)

    unmatched_stat_players = [player["Player Name"] for player in players if not player["hasStatProfile"]]
    if unmatched_stat_players:
        warnings.append("Players without any matched stat profile: " + ", ".join(unmatched_stat_players))

    for output in args.output:
        write_json(output, players)

    if warnings:
        print("PGA export warnings:", file=sys.stderr)
        for warning in warnings:
            print(f"- {warning}", file=sys.stderr)

    if args.strict_missing_stats and unmatched_stat_players:
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
