#!/usr/bin/env python3
"""Normalize the approved 16-0 workbook and 2026 schedule into one client/server dataset."""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


POSITIONS = ("QB", "RB", "WR", "TE", "K", "DST")
TEAM_ALIASES = {
    "JAC": "JAX",
    "LA": "LAR",
    "OAK": "LV",
    "SD": "LAC",
    "STL": "LAR",
    "WAS": "WSH",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument(
        "--schedule",
        default=Path("public/data/nfl/2026/games.json"),
        type=Path,
    )
    parser.add_argument(
        "--teams",
        default=Path("public/data/nfl/teams.json"),
        type=Path,
    )
    parser.add_argument(
        "--output",
        default=Path("src/features/sixteen-zero/data/players-2026-v1.json"),
        type=Path,
    )
    parser.add_argument(
        "--defense-output",
        default=Path("src/features/sixteen-zero/data/defense-ranks-2026-v1.json"),
        type=Path,
    )
    parser.add_argument("--published-at", default="2026-07-27")
    return parser.parse_args()


def clean_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def integer(value: Any) -> int | None:
    number = clean_number(value)
    return int(number) if number is not None else None


def slug(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii").lower()
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", ascii_value))


def normalize_team(value: Any) -> str | None:
    if not value:
        return None
    code = str(value).strip().upper()
    return TEAM_ALIASES.get(code, code)


def load_team_maps(teams_path: Path) -> tuple[dict[str, str], dict[str, str]]:
    payload = json.loads(teams_path.read_text(encoding="utf-8"))
    name_to_code: dict[str, str] = {}
    recognized: dict[str, str] = {}
    for team in payload["teams"]:
        code = str(team["abbr"]).upper()
        recognized[code] = str(team["name"])
        for name in {
            team.get("name"),
            team.get("fullName"),
            team.get("shortName"),
        }:
            if name:
                name_to_code[str(name).strip().lower()] = code
        name_to_code[str(team.get("nflverseAbbr", "")).strip().lower()] = code
    return name_to_code, recognized


def build_schedule(schedule_path: Path) -> dict[str, dict[int, str | None]]:
    payload = json.loads(schedule_path.read_text(encoding="utf-8"))
    schedule: dict[str, dict[int, str | None]] = defaultdict(
        lambda: {week: None for week in range(1, 18)}
    )
    for game in payload["games"]:
        if game.get("seasonType") != "REG" or not 1 <= int(game["week"]) <= 17:
            continue
        week = int(game["week"])
        home = normalize_team(game["homeAbbr"])
        away = normalize_team(game["awayAbbr"])
        if not home or not away:
            continue
        schedule[home][week] = away
        schedule[away][week] = f"@{home}"
    return dict(schedule)


def get_bye_week(team_schedule: dict[int, str | None]) -> int | None:
    missing = [week for week in range(1, 19) if team_schedule.get(week) is None]
    likely_byes = [week for week in missing if 4 <= week <= 14]
    return likely_byes[0] if likely_byes else None


def load_defense_data(
    workbook: openpyxl.Workbook,
    name_to_code: dict[str, str],
) -> tuple[
    dict[str, dict[str, int]],
    dict[str, dict[str, float]],
]:
    sheet = workbook["Points Allowed Source"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    ranks: dict[str, dict[str, int]] = {}
    points_allowed: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        name = str(row[0]).strip() if row and row[0] else ""
        code = name_to_code.get(name.lower())
        if not code:
            continue
        ranks[code] = {}
        points_allowed[code] = {}
        for position in POSITIONS:
            rank_header = f"{position} RK"
            points_header = f"{position} PA"
            if rank_header in headers:
                rank_value = integer(row[headers.index(rank_header)])
                if rank_value is not None:
                    ranks[code][position] = rank_value
            if points_header in headers:
                points_value = clean_number(row[headers.index(points_header)])
                if points_value is not None:
                    points_allowed[code][position] = points_value
    return ranks, points_allowed


def load_projection_rows(
    workbook: openpyxl.Workbook,
    name_to_code: dict[str, str],
) -> list[dict[str, Any]]:
    sheet = workbook["Projection Detail"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    result: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if not raw or raw[0] not in POSITIONS or not raw[1]:
            continue
        row = dict(zip(headers, raw))
        position = str(row["Position"])
        player_name = str(row["Player"]).strip()
        team = normalize_team(row.get("Team"))
        if position == "DST" and not team:
            team = name_to_code.get(player_name.lower())
        projected = clean_number(row.get("Recalculated PPR FPTS"))
        if projected is None:
            projected = clean_number(row.get("FantasyPros FPTS"))
        result.append(
            {
                "position": position,
                "name": player_name,
                "team": team,
                "projectedSeasonPoints": projected,
            }
        )
    return result


def load_ranked_rows(workbook: openpyxl.Workbook) -> list[dict[str, Any]]:
    sheet = workbook["Rankings V1"]
    rows = list(sheet.iter_rows(values_only=True))
    headers = list(rows[0])
    result: list[dict[str, Any]] = []
    for raw in rows[1:]:
        row = dict(zip(headers, raw))
        if row.get("Pos") not in POSITIONS or not row.get("Player"):
            continue
        position = str(row["Pos"])
        projected = clean_number(row.get("FantasyPros PPR Pts"))
        blended = clean_number(row.get("Blended Projection"))
        consensus_implied = clean_number(row.get("Consensus-Implied Pts"))
        approved_fallback = blended or consensus_implied or projected
        result.append(
            {
                "name": str(row["Player"]).strip(),
                "team": normalize_team(row.get("Team")),
                "position": position,
                "byeWeekWorkbook": integer(row.get("Bye")),
                "consensusOverallRank": integer(row.get("Sharp Rank")),
                "consensusPositionRank": integer(
                    re.sub(r"[^0-9]", "", str(row.get("Sharp Pos Rank") or ""))
                ),
                "projectedSeasonPoints": projected or approved_fallback,
                "projectionPositionRank": integer(row.get("FP Pos Rank")),
                "blendedSeasonPoints": blended or approved_fallback,
                "fullSeasonSOSRank": integer(row.get("SOS")),
                "usedProjectionFallback": projected is None,
            }
        )
    return result


def add_special_teams(
    ranked: list[dict[str, Any]],
    projections: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    additions: list[dict[str, Any]] = []
    for position in ("K", "DST"):
        eligible = [
            row
            for row in projections
            if row["position"] == position
            and row["team"]
            and (row["projectedSeasonPoints"] or 0) > 0
        ]
        eligible.sort(
            key=lambda row: (
                -(row["projectedSeasonPoints"] or 0),
                row["name"],
            )
        )
        for position_rank, row in enumerate(eligible[:32], start=1):
            additions.append(
                {
                    **row,
                    "byeWeekWorkbook": None,
                    "consensusOverallRank": None,
                    "consensusPositionRank": position_rank,
                    "projectionPositionRank": position_rank,
                    "blendedSeasonPoints": row["projectedSeasonPoints"],
                    "fullSeasonSOSRank": None,
                    "usedProjectionFallback": False,
                }
            )
    additions.sort(
        key=lambda row: (
            -(row["blendedSeasonPoints"] or 0),
            row["position"],
            row["name"],
        )
    )
    next_rank = len(ranked) + 1
    for row in additions:
        row["consensusOverallRank"] = next_rank
        next_rank += 1
    return ranked + additions


def opponent_code(opponent: str | None) -> str | None:
    return opponent.replace("@", "") if opponent else None


def assign_position_ranks(rows: list[dict[str, Any]], value_key: str, output_key: str) -> None:
    by_position: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_position[row["position"]].append(row)
    for position_rows in by_position.values():
        position_rows.sort(
            key=lambda row: (
                -(row.get(value_key) or 0),
                row["consensusOverallRank"],
            )
        )
        for rank, row in enumerate(position_rows, start=1):
            row[output_key] = rank


def assign_playoff_sos(
    rows: list[dict[str, Any]],
    schedules: dict[str, dict[int, str | None]],
    defense_ranks: dict[str, dict[str, int]],
) -> None:
    by_position: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        ranks = []
        for week in (15, 16, 17):
            opponent = opponent_code(schedules.get(row["team"], {}).get(week))
            rank = defense_ranks.get(opponent or "", {}).get(row["position"])
            if rank is not None:
                ranks.append(rank)
        row["_playoffSosMean"] = sum(ranks) / len(ranks) if ranks else None
        by_position[row["position"]].append(row)
    for position_rows in by_position.values():
        ranked = sorted(
            [row for row in position_rows if row["_playoffSosMean"] is not None],
            key=lambda row: (row["_playoffSosMean"], row["consensusOverallRank"]),
        )
        for rank, row in enumerate(ranked, start=1):
            row["playoffSOSRank"] = rank


def normalize_players(
    rows: list[dict[str, Any]],
    schedules: dict[str, dict[int, str | None]],
    defense_ranks: dict[str, dict[str, int]],
    points_allowed: dict[str, dict[str, float]],
    recognized_teams: dict[str, str],
) -> list[dict[str, Any]]:
    assign_position_ranks(rows, "blendedSeasonPoints", "blendedPositionRank")
    for row in rows:
        if not row.get("projectionPositionRank"):
            row["projectionPositionRank"] = row["blendedPositionRank"]
    assign_playoff_sos(rows, schedules, defense_ranks)

    players: list[dict[str, Any]] = []
    for row in sorted(rows, key=lambda item: item["consensusOverallRank"]):
        team = row["team"]
        if team not in recognized_teams:
            raise ValueError(f"Unrecognized team code for {row['name']}: {team}")
        team_schedule = schedules.get(team, {})
        weekly_opponents = {
            str(week): team_schedule.get(week) for week in range(1, 18)
        }
        opponent_allowed = {}
        for week in range(1, 18):
            opponent = opponent_code(team_schedule.get(week))
            opponent_allowed[str(week)] = (
                points_allowed.get(opponent or "", {}).get(row["position"])
                if opponent
                else None
            )
        bye_week = get_bye_week(team_schedule) or row.get("byeWeekWorkbook")
        projected = round(float(row["projectedSeasonPoints"]), 3)
        blended = round(float(row["blendedSeasonPoints"]), 3)
        completeness_checks = [
            team,
            bye_week,
            projected,
            blended,
            row.get("fullSeasonSOSRank"),
            row.get("playoffSOSRank"),
            sum(value is not None for value in weekly_opponents.values()) >= 16,
            sum(value is not None for value in opponent_allowed.values()) >= 16,
        ]
        players.append(
            {
                "id": f"{row['position'].lower()}-{slug(row['name'])}-{team.lower()}",
                "name": row["name"],
                "team": team,
                "position": row["position"],
                "byeWeek": bye_week,
                "consensusOverallRank": row["consensusOverallRank"],
                "consensusPositionRank": row["consensusPositionRank"],
                "projectedSeasonPoints": projected,
                "projectedPPG": round(projected / 17, 3),
                "projectionPositionRank": row["projectionPositionRank"],
                "blendedSeasonPoints": blended,
                "blendedPPG": round(blended / 17, 3),
                "blendedPositionRank": row["blendedPositionRank"],
                "fullSeasonSOSRank": row.get("fullSeasonSOSRank"),
                "playoffSOSRank": row.get("playoffSOSRank"),
                "weeklyOpponents": weekly_opponents,
                "opponentFantasyPointsAllowed": opponent_allowed,
                "dataCompleteness": round(
                    sum(bool(value) for value in completeness_checks)
                    / len(completeness_checks),
                    3,
                ),
                "active": True,
            }
        )
    return players


def validate(players: list[dict[str, Any]]) -> dict[str, Any]:
    errors: list[str] = []
    if len(players) < 275:
        errors.append(f"Expected at least 275 active players; found {len(players)}.")
    ids = [player["id"] for player in players]
    if len(ids) != len(set(ids)):
        errors.append("Player IDs are not unique.")
    identities = [(player["name"].casefold(), player["position"]) for player in players]
    if len(identities) != len(set(identities)):
        errors.append("Duplicate player-position records exist.")
    invalid_positions = sorted(
        {player["position"] for player in players if player["position"] not in POSITIONS}
    )
    if invalid_positions:
        errors.append(f"Invalid positions: {', '.join(invalid_positions)}")
    without_projection = [
        player["name"]
        for player in players
        if not player["projectedSeasonPoints"] or not player["blendedSeasonPoints"]
    ]
    if without_projection:
        errors.append(f"Players without approved projections: {', '.join(without_projection)}")
    counts = {
        position: sum(player["position"] == position for player in players)
        for position in POSITIONS
    }
    if counts["K"] < 24 or counts["DST"] < 24:
        errors.append("Player universe cannot supply two K and two DST to all 12 teams.")
    if errors:
        raise ValueError("\n".join(errors))
    return {
        "activePlayers": len(players),
        "positionCounts": counts,
        "uniqueIds": len(set(ids)),
        "minimumCompleteness": min(player["dataCompleteness"] for player in players),
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    args = parse_args()
    workbook = openpyxl.load_workbook(args.workbook, data_only=True, read_only=True)
    name_to_code, recognized_teams = load_team_maps(args.teams)
    schedules = build_schedule(args.schedule)
    defense_ranks, points_allowed = load_defense_data(workbook, name_to_code)
    projections = load_projection_rows(workbook, name_to_code)
    ranked = load_ranked_rows(workbook)
    combined = add_special_teams(ranked, projections)
    excluded_unavailable = [
        row["name"] for row in combined if row.get("team") not in recognized_teams
    ]
    combined = [
        row for row in combined if row.get("team") in recognized_teams
    ]
    players = normalize_players(
        combined,
        schedules,
        defense_ranks,
        points_allowed,
        recognized_teams,
    )
    validation = validate(players)
    workbook_modified = datetime.fromtimestamp(
        args.workbook.stat().st_mtime,
        tz=timezone.utc,
    ).isoformat()
    write_json(
        args.output,
        {
            "_meta": {
                "dataVersion": "fantasy-2026-v2",
                "season": 2026,
                "publishedAt": args.published_at,
                "workbook": args.workbook.name,
                "workbookModifiedAt": workbook_modified,
                "scheduleSource": str(args.schedule).replace("\\", "/"),
                "projectionGames": 17,
                "excludedUnavailablePlayers": excluded_unavailable,
                "validation": validation,
            },
            "players": players,
        },
    )
    write_json(
        args.defense_output,
        {
            "_meta": {
                "dataVersion": "fantasy-2026-v2",
                "season": 2026,
                "publishedAt": args.published_at,
                "interpretation": "Rank 1 allowed the most fantasy points; rank 32 allowed the least.",
            },
            "defenseRanks": defense_ranks,
        },
    )
    print(json.dumps(validation, indent=2))


if __name__ == "__main__":
    main()
